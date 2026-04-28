import os
from flask import Flask, send_from_directory, request, jsonify
import openpyxl # Para ver e editar arquivos .xlsx
from datetime import (
    datetime,
)
# Caminho base do projeto (uma pastacima do backend)
BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))

# Pasta frontend (HTML, JS)
FRONTEND_DIR = os.path.join(BASE_DIR, "frontend")

# Pasta static (CSS)
STATIC_DIR = os.path.join(BASE_DIR, "static")

# Pasta db (Banco de Dados)
DB_DIR = os.path.join(BASE_DIR, "db")
EXCEL_FILE = os.path.join(DB_DIR, "clientes.xlsx")

# Pasta do JS
JS_DIR = os.path.join(BASE_DIR, "frontend", "js")
JS_FILE = os.path.join(JS_DIR, "main.js")

# Cabecalhos das colunas do Excel (linha 1)
COLUNAS = [
    "ID",
    "Nome",
    "CPF",
    "Email",
    "Telefone",
    "Endereço",
    "Observações",
    "Data Cadastro"
]

def init_excel():
    if not os.path.exists(DB_DIR):
        os.makedirs(DB_DIR) # Cria a pasta se não existir
    
    if not os.path.exists(EXCEL_FILE):
        workbook = openpyxl.Workbook() # Cria a planilha
        sheet = workbook.active # Pega a planilha ativa
        sheet.title = "Clientes" # Nomeia a aba principal
        sheet.append(COLUNAS) # Adiciona os títulos das colunas
        workbook.save(EXCEL_FILE) # Salva o arquivo Excel

app = Flask(__name__, static_folder=STATIC_DIR, static_url_path="/static")

# Pagina principal
@app.route("/")
def home():
    return send_from_directory(FRONTEND_DIR, "index.html")

# Pagina de consulta
@app.route("/consulta")
def consulta_page():
    return send_from_directory(FRONTEND_DIR, "consultar.html")

# Pagina de alteracao
@app.route("/alterar")
def alterar_page():
    return send_from_directory(FRONTEND_DIR, "alterar.html")

# Rota para servir imagens, sripts ou outros arquivos na pasta "assets"
@app.route("/assets/<path:filename>")
def assets(filename):
    return send_from_directory("../frontend/assets", filename)

# INFOS PROFESSOR
# Nome: José Antônio dos Reis
# CPF: 087.280.297.39
# Email: apoio.klesis@gmail.com
# Telefone: 41985147636
# Rua das Nações Unidas, 871

# -------------------------------------------------------------------------------------------------
# CADASTRAR CLIENTE
# -------------------------------------------------------------------------------------------------
@app.route("/cadastrar", methods=["POST"])
def cadastrar_cliente():
    """
    Recebe os dados do formulário (em JSON), valida e salva um novo cliente
    """
    try:
        data = request.json # Dados enviados do frontend via POST (JSON)

        # Capmos obrigatórios que o usuário deve preencher
        required_fields = ["nome", "cpf", "email", "telefone", "endereco"]
        if not all(field in data and data[field] for field in required_fields):
            return (
                jsonify(
                    {
                        "status": "error",
                        "message": "Todos os campos obrigatórios devem ser preenchidos."
                    }
                ),
                400,
            )
        workbook = openpyxl.load_workbook(EXCEL_FILE) # Abre o arquivo Excel
        sheet = workbook.active

        # Cria um ID automático (último ID + 1)
        last_id = 0
        if sheet.max_row > 1:
            last_id = sheet.cell(row=sheet.max_row, column=1).value or 0
        new_id = last_id + 1

        # Cria uma nova linha com os dados informados
        novo_cliente = [
            new_id,
            data.get("nome"),
            data.get("cpf"),
            data.get("email"),
            data.get("telefone"),
            data.get("endereco"),
            data.get("observacoes", ""), # Campo opcional
            datetime.now().strftime("%Y-%m-%d") # Data atual
        ]

        sheet.append(novo_cliente) # Adiciona nova linha no Excel
        workbook.save(EXCEL_FILE) # Salva alterações

        # Retorna mensagem de sucesso
        return (
            jsonify(
                {
                "status": "sucess",
                "message": "Cliente cadastrado com sucesso!",
                "id": new_id
                }
            ),
            201,
        )
    except Exception as e:
        # Tratamento de erro genérico
        return (
            jsonify(
                {
                "status": "error",
                "message": f"Erro ao salvar no servidor: {e}"
                }
            ),
            500,
        )

# -------------------------------------------------------------------------------------------------
# CONSULTAR CLIENTE PELO NOME
# -------------------------------------------------------------------------------------------------

@app.route("/buscar", methods=["GET"])
def buscar_cliente():
    """
    Busca clientes pelo nome (não diferencia maiúsculas/minúsculas)
    """
    nome_query = request.args.get("nome", "").lower() # Nome pesquisado

    try:
        workbook = openpyxl.load_workbook(EXCEL_FILE) # Abre o arquivo Excel
        sheet = workbook.active
        resultados = [] # Lista para armazenar clientes encontrados

        # Percorre as linhas do Excel (começando da linha 2, pulando os títulos)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            cliente = dict(zip(COLUNAS, row)) # Cria um dicionário com os dados do cliente
            nome_cliente = (cliente.get("Nome") or "").lower() # Nome do cliente em minúsculas

            if nome_query in nome_cliente:
                resultados.append(cliente)

        return jsonify(resultados) # Retorna a lista de clientes encontrados
    
    except FileNotFoundError:
        return (
            jsonify(
                {
                    "status": "error",
                    "message": "Arquivo de dados não encontrado."
                }
            ),
            404,
        )
    except Exception as e:
        return (
            jsonify(
                {
                    "status": "error",
                    "message": f"Erro ao ler os dados: {e}"
                }
            ),
            500,
        )

# -------------------------------------------------------------------------------------------------
# CONSULTAR CLIENTE PELO ID
# -------------------------------------------------------------------------------------------------

@app.route("/api/cliente/<int:cliente_id>", methods=["GET"])
def get_cliente(cliente_id):
    """
    Retorna os dados completos de um cliente pelo seu ID
    """
    try:
        workbook = openpyxl.load_workbook(EXCEL_FILE) # Abre o arquivo Excel
        sheet = workbook.active

        # Procura o cliente linha por linha
        for row_idx in range(2, sheet.max_row + 1):
            row_id = sheet.cell(row=row_idx, column=1).value
            if row_id == cliente_id:
                row_values = [cell.value for cell in sheet[row_idx]]
                cliente = dict(zip(COLUNAS, row_values))
                return jsonify(cliente)

        # Se não encontrar o cliente
        return jsonify(
            {
                "status": "error",
                "message": "Cliente não encontrado."
            }
        )

    except Exception as e:
        return (
            jsonify(
                {
                    "status": "error",
                    "message": f"Erro ao buscar clientes: {e}"
                }
            ),
            500,
        )

# -------------------------------------------------------------------------------------------------
# ALTERAR OS DADOS DO CLIENTE
# -------------------------------------------------------------------------------------------------

@app.route("/api/alterar/<int:cliente_id>", methods=["POST"])
def alterar_cliente(cliente_id):
    """
    Recebe os dados atualizados de um cliente e salva as alterações
    """
    try:
        data = request.json # Dados enviados do frontend via POST (JSON)

        workbook = openpyxl.load_workbook(EXCEL_FILE) # Abre o arquivo Excel
        sheet = workbook.active

        # Comecamos com -1 para indicar que, por enquanto, nao o encontramos
        row_to_update = -1

        # O sistema comeca a let da linha 2 (pulando os titulos) ate a ultima linha
        for row_idx in range(2, sheet.max_row + 1):
            # Se o valor da primeira coluna (ID) for igual ao id que recebemos (cliente_id), achamos!
            if sheet.cell(row=row_idx, column=1).value == cliente_id:
                row_to_update = row_idx
                break

        # Se depois de ler tudo, continuarmos com -1, o hospede nao existe
        if row_to_update == -1:
            return (
                jsonify(
                    {
                        "status": "error",
                        "message": "Cliente não encontrado."
                    }
                ),
                404, # Codigo de erro padrao para "Nao Encontrado"
            )

        # coluna 2 = Nome, coluna 3 = CPF, coluna 4 = Email, coluna 5 = Telefone, coluna 6 = Endereco, coluna 7 = Observacoes
        sheet.cell(row=row_to_update, column=2).value = data.get("nome") # Nome
        sheet.cell(row=row_to_update, column=3).value = data.get("cpf") # CPF
        sheet.cell(row=row_to_update, column=4).value = data.get("email") # Email
        sheet.cell(row=row_to_update, column=5).value = data.get("telefone") # Telefone
        sheet.cell(row=row_to_update, column=6).value = data.get("endereco") # Endereco
        sheet.cell(row=row_to_update, column=7).value = data.get("observacoes") # Observacoes

        # IMPORTANTE: Sempre que fizer alterações no Excel, é necessário salvar o arquivo para que as mudanças sejam efetivadas
        workbook.save(EXCEL_FILE)

        # Retorna mensagem de sucesso para o front-end mostrar na tela do usuario
        return jsonify(
            {
                "status": "success",
                "message": "Dados do cliente atualizado com sucesso!"
            }
        )
    
    except Exception as e:
        # Se houver algum erro inesperado (ex: o arquivo e=Excel estar aberto por outro programa)
        return (
            jsonify(
                {
                    "status": "error",
                    "message": f"Erro ao atualizar os dados: {e}"
                }
            ),
            500, # Codigo de erro para "Erro Interno do Servidor"
        )

if __name__ == "__main__":
    # print("Base: ", BASE_DIR)
    # print("Front: ", FRONTEND_DIR)
    # print("Static:", STATIC_DIR)
    # print("Excel:", EXCEL_FILE)
    # print("JS:", JS_FILE)
    init_excel()
    app.run(debug=False)